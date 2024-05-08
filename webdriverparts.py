from selenium import webdriver
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import os
import sys

num = int(sys.argv[1]) if len(sys.argv) > 1 else 0

# ここで num を使用して何かを実行する
(f"受け取った引数: {num}")
if num <= 1:
    crl_p = "crl"
else:
    crl_p = "crl_p" + str(num)

url = f"https://watchnian.com/shop/c/{crl_p}/?filtercode13=1#block_of_filter"


def prices_array_make(logs):
    prices_array = []
    prices_array_item = []
    itemindex = 1
    for index, log in enumerate(logs):
        if itemindex % 2 == 0:
            prices_array_item.append(log)
            prices_array.append(prices_array_item)
            prices_array_item = []
            itemindex = 1
        else:
            prices_array_item.append(log)
            itemindex = 2
    return prices_array


# refを取得する
def ref_array_make(logs):
    itemarray = []
    for index, log in enumerate(logs):
        item = str(log[0]) + str(log[1])
        itemarray.append(str(item))
    save_logs_to_file(itemarray, "ref.txt")
    return itemarray


# 色と製品番号を取得する
def color_array_make(logs):
    # 最初の値を取得するための正規表現
    pattern = r"\[([^\]]+)\]お気に入り登録"
    parentheses = r"\((.*?)\)お気に入り登録"
    itemarray = []
    # logs = extract_brackets(logs)

    for index, log in enumerate(logs):
        itemarray.append(str(log))
        # itemarray.append(log)
        if len(log) > 2:
            matches = re.search(pattern, log[1])
            parenthesesmatches = re.search(parentheses, log[0])
            if matches:
                # itemarray.append(matches.group(1))
                continue

            if parenthesesmatches:
                itemarray.append(parenthesesmatches.group(0))
                itemarray.append(parenthesesmatches.group(1))
                itemarray.append(parenthesesmatches.group(2))
                itemarray.append(parenthesesmatches.group(3))
                continue
    combined_list = list(map(lambda x: x[0] + x[1], itemarray))

    save_logs_to_file(itemarray, "color.txt")
    return itemarray


# 正規表現で取得したリファレンスナンバーと色を抽出しておなじ配列にする関数
def refandcollor_array_make(logs):
    itemarray = []
    for log in logs:
        (log[0] if log[0] else log[1])
        itemarray.append(str(log[0]) if log[0] else str(log[1]))
    save_logs_to_file(itemarray, "refandcolor.txt")


def textprocess(text):
    # お気に入り登録以降を削除する正規表現
    pattern = re.compile(r"お気に入り登録\d+.*?$", flags=re.DOTALL)
    # テキストを加工してお気に入り登録以降を削除
    processed_text = re.sub(pattern, "", text, flags=re.DOTALL)

    # 結果を表示
    (processed_text)


def extract_brackets(text):
    # '数字件' パターンを削除
    text = [
        re.sub(r".*?(\w+\[.*?件.*?\]|\w+\(.*?件.*?\))", "", str(item)) for item in text
    ]
    # text = re.sub(r'\d+件', '', text)
    # 正規表現パターン
    pattern = r"\[([^\]]+)\]|\((.*?)\)"
    matches = [re.findall(pattern, item) for item in text]

    text = "\n".join(map(str, text))
    #  マッチング
    matches = re.findall(pattern, text)
    matches = [item for item in matches if "件" not in item]
    # 結果を返す
    return matches


def textlog(text, file_path="text.txt"):
    with open("greentext.txt", "w", encoding="utf_8") as file:

        for item in text:
            # parenthesesvaluematches = re.findall(r"\[([^]]+)\]", tbody_text)
            color_in_parentheses = re.findall(r"\[([^]]+)\]", str(item))
            pattern = r"\[([^\]]+)\]お気に入り登録"
            color_in_parentheses = re.findall(r"\[([^\]]+)\]お気に入り登録", str(item))
            matches = re.search(pattern, item)

            # 括弧内の文字列を取得
            matches_in_parentheses = re.findall(r"\((.*?)\)", str(item))

            # パターンにマッチする部分を取得
            matches_pattern = re.findall(r"\b(\d{4,6})([a_zA_Z]+)?", str(item))

            # ファイルに書き込み
            if matches:
                file.write("Matcheskako" + ",".join(matches) + "\n")
            if color_in_parentheses:
                file.write("Matcheskako" + ",".join(matches) + "\n")
            if matches_in_parentheses:
                file.write(
                    "Matches in parentheses: "
                    + ", ".join(matches_in_parentheses)
                    + "\n"
                )
            if matches_pattern:
                file.write(
                    "Matches pattern: "
                    + ", ".join(["".join(match) for match in matches_pattern])
                    + "\n"
                )
            file.write("____________\n")

    with open(file_path, "w", encoding="utf_8") as file:
        if isinstance(text, list):
            text = [str(item) + "\n____________" for item in text]

            text = "\n".join(map(str, text))
            file.write(text)
            file.write("____________")

        else:
            file.write(text)
    return text


# ログをファイルに保存する関数
def save_logs_to_file(logs, file_path):
    # ここでアイテム一覧の配列を作ってしまう。
    # ここでの配列は二つで一つの二次元配列になる。
    with open(file_path, "w", encoding="utf_8") as file:
        i = 1
        for index, log in enumerate(logs):
            if i % 2 == 0:
                file.write(str(log) + "はデータです" + "\n")
                file.write("________________" + "\n")
                i = 1
            else:
                file.write(str(log) + "はデータです" + "\n")
                i = 2


def validate_input(input_string):
    pattern = re.compile(r"^\d{4,10}[a_zA_Z]*$")
    return bool(pattern.match(input_string))


# 金額抽出
def price_validate_imput(input_string):
    pattern = re.compile(r"￥(\d+)")
    return bool(pattern.match(input_string))


def dltag_getitem(item_soup, item):
    pricrice_tag = item_soup.find("dt", text=item)
    price_value_tag = pricrice_tag.find_next_sibling("dd")
    print(price_value_tag)
    print(item)
    print("map関数")


# エクセルのヘッダ－データ
data = ["商品番号", "モデル", "最高", "その他"]
# この中を各辞書型にする。
itemlist = []

# 金額配列
pricelist = []

# アイテム名リスト
itemnamelist = []


# 現在の日付を取得
today_date = datetime.now().strftime("%Y%m%d")
# ファイル名に日付を組み込む
file_name = f"output_{today_date}.xlsx"
if not os.path.exists(file_name):
    # Excelブックの作成
    wb = Workbook()
    ws = wb.active
    # ヘッダー行を追加
    ws.append(["モデル名", "リファレンスNO", "文字盤", "ブレスレット", "価格", "URL"])
else:
    # ファイルが存在する場合は既存のファイルを読み込み
    wb = load_workbook(file_name)
    ws = wb.active
# SeleniumのWebDriverを初期化
driver = webdriver.Chrome()  # または他のブラウザに合わせて選択

# URLを開く
driver.get(url)

# Seleniumがページのロードを待つなどの適切な待機処理が必要な場合はここで実施

# ページのHTMLを取得
page_source = driver.page_source

# BeautifulSoupを使ってHTMLを解析
soup = BeautifulSoup(page_source, "html.parser")

# <tbody> タグ内のテキストを抽出して表示
tbody_tag = soup.find("body")

# <main>タグを抽出
main_tag = tbody_tag.find("main")
# 各アイテムを抽出する。
# <main>のクラスを抽出ここで各アイテムの値段を抽出する。
topcontents_newarrival_slide = main_tag.find_all(class_="topcontents_newarrival_slide")
blockthumnail = main_tag.find(class_="block-category-list")

# !テスト変数
blockthumnail = main_tag.find(class_="block-category-list--goods")

goos = blockthumnail.find_all(class_="block-thumbnail-t--goods-name")

# topcontents_newarrival_slide = tbody_tag.find(class_="MarketSearch2")
# 以下でliタグをループする。
# print(topcontents_newarrival_slide)
# 時計一覧
row_item = []
for element in goos:
    # print(element)
    # url取得する。
    modelname = ""
    priceget = ""
    refarence_no_value_tag = ""
    bracelet_value_tag = ""
    display_value_tag = ""
    url_joint = ""

    atug = element.find("a")
    href_value = atug.get("href")
    # /shop/g/gik-00-0553155/

    url_joint = "https://watchnian.com" + href_value
    print(url_joint)
    driver.get(url_joint)

    item_page_source = driver.page_source
    item_soup = BeautifulSoup(item_page_source, "html.parser")
    bodytag = item_soup.find("body")

    # 欲しいやつを配列にして、map関数にする
    element_categoris = [
        "モデル名",
        "リファレンスNO",
        "文字盤",
        "ブレスレット",
        "価格",
        "URL",
    ]

    #!謎のタグ取得不可エラーが発生した際に利用する。
    # modelnameget = item_soup.find("dt", string="モデル")
    # modelname = modelnameget.find_next_sibling("dd").get_text(strip=True)

    # モデル名
    try:
        modelnameget = item_soup.find("dt", string="モデル")
        modelname = modelnameget.find_next_sibling("dd").get_text(strip=True)
    except AttributeError:
        print("モデル名なし")

    # input_values = list(map(dltag_getitem, item_soup, element_categoris))
    try:
        reference_no_tag = item_soup.find("dt", string="型番（型式番号）")
        refarence_no_value_tag = reference_no_tag.find_next_sibling("dd").get_text(
            strip=True
        )
    except AttributeError:
        print("リファレンスナンバーがなかった")

    # 金額
    try:
        priceget = item_soup.find(class_="price_body").get_text(strip=True)
        # pricrice_tag = item_soup.find("dt", text="定価")
        # price_value_tag = pricrice_tag.find_next_sibling("dd").get_text(strip=True)
        display_tag = item_soup.find("dt", string="文字盤")
        display_value_tag = display_tag.find_next_sibling("dd").get_text(strip=True)
    except AttributeError:
        print("文字盤がない")

    try:
        bracelet_tag = item_soup.find("dt", string="ブレスレット")
        bracelet_value_tag = bracelet_tag.find_next_sibling("dd").get_text(strip=True)
    except AttributeError:
        print("ブレスレットがない")
    # モデル名
    row_item.append(modelname)

    # リファレンスナンバー
    row_item.append(refarence_no_value_tag)

    # 文字盤
    row_item.append(display_value_tag)

    # ブレスレット
    row_item.append(bracelet_value_tag)

    # 金額
    row_item.append(priceget)

    # URL
    row_item.append(url_joint)
    # dltag = bodytag.find_all("dl")
    # print(dltag)
    # target_pricetag = dltag.find("dt", text="定価")
    # print(target_pricetag)
    print("-------------------")
    ws.append(row_item)
    row_item = []
wb.save(file_name)
# a_tag_get = main_tag.find_all(class_="js-enhanced-ecommerce-image")

for element in topcontents_newarrival_slide:
    # !urlを取得しなければいけない
    # !aタグを取得して、そこからまた通信する
    # !js-enhanced-ecommerce-image クラスを取得して href を取得
    # ここのループでデータを抽出する。値段、リファレンスナンバー etc....
    # 値段、REF、ブレスレット
    # aタグを抽出する。

    # print(element)
    blockthumbnailtgoodsname = element.find(class_="block-thumbnail-t--goods-name")
    atug = blockthumbnailtgoodsname.find("a", class_="js-enhanced-ecommerce-goods-name")
    if atug:
        testcategori = atug.get("data-category3")
        # print("data-category3の値:", testcategori)

    category_text = blockthumbnailtgoodsname.get_text(strip=True)
    # 空白分割でアイテム一覧を取得する。
    category_text_words = category_text.split()
    # print(category_text_words)
    # 値段
    price = element.find(class_="price_body")
    print(price)
    print("------------------------------")
