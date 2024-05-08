from selenium import webdriver
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import os
import sys

num = int(sys.argv[1]) if len(sys.argv) > 1 else 0

# ここで num を使用して何かを実行する
(f"受け取った引数: {num}")
if num <= 1:
    crl_p = "crl"
else:
    crl_p = "crl_p" + num

url = f"https://watchnian.com/shop/c/{crl_p}/?filtercode13=1#block_of_filter"


def prices_array_make(logs):
    prices_array = []
    prices_array_item = []
    itemindex = 1
    for index, log in enumerate(logs):
        if itemindex % 2 == 0:
            prices_array_item.append(log)
            prices_array.append(prices_array_item)
            prices_array_item = []
            itemindex = 1
        else:
            prices_array_item.append(log)
            itemindex = 2
    return prices_array


# refを取得する
def ref_array_make(logs):
    itemarray = []
    for index, log in enumerate(logs):
        item = str(log[0]) + str(log[1])
        itemarray.append(str(item))
    save_logs_to_file(itemarray, "ref.txt")
    return itemarray


# 色と製品番号を取得する
def color_array_make(logs):
    # 最初の値を取得するための正規表現
    pattern = r"\[([^\]]+)\]お気に入り登録"
    parentheses = r"\((.*?)\)お気に入り登録"
    itemarray = []
    # logs = extract_brackets(logs)

    for index, log in enumerate(logs):
        itemarray.append(str(log))
        # itemarray.append(log)
        if len(log) > 2:
            matches = re.search(pattern, log[1])
            parenthesesmatches = re.search(parentheses, log[0])
            if matches:
                # itemarray.append(matches.group(1))
                continue

            if parenthesesmatches:
                itemarray.append(parenthesesmatches.group(0))
                itemarray.append(parenthesesmatches.group(1))
                itemarray.append(parenthesesmatches.group(2))
                itemarray.append(parenthesesmatches.group(3))
                continue
    combined_list = list(map(lambda x: x[0] + x[1], itemarray))

    save_logs_to_file(itemarray, "color.txt")
    return itemarray


# 正規表現で取得したリファレンスナンバーと色を抽出しておなじ配列にする関数
def refandcollor_array_make(logs):
    itemarray = []
    for log in logs:
        (log[0] if log[0] else log[1])
        itemarray.append(str(log[0]) if log[0] else str(log[1]))
    save_logs_to_file(itemarray, "refandcolor.txt")


def textprocess(text):
    # お気に入り登録以降を削除する正規表現
    pattern = re.compile(r"お気に入り登録\d+.*?$", flags=re.DOTALL)
    # テキストを加工してお気に入り登録以降を削除
    processed_text = re.sub(pattern, "", text, flags=re.DOTALL)

    # 結果を表示
    (processed_text)


def extract_brackets(text):
    # '数字件' パターンを削除
    text = [
        re.sub(r".*?(\w+\[.*?件.*?\]|\w+\(.*?件.*?\))", "", str(item)) for item in text
    ]
    # text = re.sub(r'\d+件', '', text)
    # 正規表現パターン
    pattern = r"\[([^\]]+)\]|\((.*?)\)"
    matches = [re.findall(pattern, item) for item in text]

    text = "\n".join(map(str, text))
    #  マッチング
    matches = re.findall(pattern, text)
    matches = [item for item in matches if "件" not in item]
    # 結果を返す
    return matches


def textlog(text, file_path="text.txt"):
    with open("greentext.txt", "w", encoding="utf-8") as file:

        for item in text:
            # parenthesesvaluematches = re.findall(r"\[([^]]+)\]", tbody_text)
            color_in_parentheses = re.findall(r"\[([^]]+)\]", str(item))
            pattern = r"\[([^\]]+)\]お気に入り登録"
            color_in_parentheses = re.findall(r"\[([^\]]+)\]お気に入り登録", str(item))
            matches = re.search(pattern, item)

            # 括弧内の文字列を取得
            matches_in_parentheses = re.findall(r"\((.*?)\)", str(item))

            # パターンにマッチする部分を取得
            matches_pattern = re.findall(r"\b(\d{4,6})([a-zA-Z]+)?", str(item))

            # ファイルに書き込み
            if matches:
                file.write("Matcheskako" + ",".join(matches) + "\n")
            if color_in_parentheses:
                file.write("Matcheskako" + ",".join(matches) + "\n")
            if matches_in_parentheses:
                file.write(
                    "Matches in parentheses: "
                    + ", ".join(matches_in_parentheses)
                    + "\n"
                )
            if matches_pattern:
                file.write(
                    "Matches pattern: "
                    + ", ".join(["".join(match) for match in matches_pattern])
                    + "\n"
                )
            file.write("------------\n")

    with open(file_path, "w", encoding="utf-8") as file:
        if isinstance(text, list):
            text = [str(item) + "\n------------" for item in text]

            text = "\n".join(map(str, text))
            file.write(text)
            file.write("------------")

        else:
            file.write(text)
    return text


# ログをファイルに保存する関数
def save_logs_to_file(logs, file_path):
    # ここでアイテム一覧の配列を作ってしまう。
    # ここでの配列は二つで一つの二次元配列になる。
    with open(file_path, "w", encoding="utf-8") as file:
        i = 1
        for index, log in enumerate(logs):
            if i % 2 == 0:
                file.write(str(log) + "はデータです" + "\n")
                file.write("----------------" + "\n")
                i = 1
            else:
                file.write(str(log) + "はデータです" + "\n")
                i = 2


def validate_input(input_string):
    pattern = re.compile(r"^\d{4,10}[a-zA-Z]*$")
    return bool(pattern.match(input_string))


# 金額抽出
def price_validate_imput(input_string):
    pattern = re.compile(r"￥(\d+)")
    return bool(pattern.match(input_string))


# エクセルのヘッダ－データ
data = ["商品番号", "モデル", "最高", "その他"]
# この中を各辞書型にする。
itemlist = []

# 金額配列
pricelist = []

# アイテム名リスト
itemnamelist = []


# 現在の日付を取得
today_date = datetime.now().strftime("%Y%m%d")
# ファイル名に日付を組み込む

file_name = f"output_{today_date}.xlsx"
if not os.path.exists(file_name):
    # Excelブックの作成
    wb = Workbook()
    ws = wb.active
    # ヘッダー行を追加
    ws.append(
        ["製品名", "リファレンスNO", "最高価格", "最安価格", "ブレスレット", "その他"]
    )
else:
    # ファイルが存在する場合は既存のファイルを読み込み
    wb = load_workbook(file_name)
    ws = wb.active
# SeleniumのWebDriverを初期化
driver = webdriver.Chrome()  # または他のブラウザに合わせて選択

# URLを開く
driver.get(url)

# Seleniumがページのロードを待つなどの適切な待機処理が必要な場合はここで実施

# ページのHTMLを取得
page_source = driver.page_source


# BeautifulSoupを使ってHTMLを解析
soup = BeautifulSoup(page_source, "html.parser")


# <tbody> タグ内のテキストを抽出して表示
tbody_tag = soup.find("body")
print(tbody_tag)


if tbody_tag:
    tbody_text = tbody_tag.get_text(strip=True)
    # 空白で分割する
    text_words = tbody_text.split()

    textlog = textlog(text_words)

    # textlog = textlog(tbody_text)
    # textprocess(textlog)
    # 正規表現で[]で囲まれたテキストを抽出する
    parenthesesvaluematches = re.findall(r"\[([^]]+)\]", tbody_text)
    kako = re.findall(r"\([^\)]*\)", tbody_text)
    pricevaluematches = re.findall(r"￥(\d+)", tbody_text)
    # refalenceNoと色の正規表現
    refandcolormatches = re.findall(
        # r"\d{4,10}[a-zA-Z]*(?:\[([^]]+)\]|\([^\)]*\))", tbody_text
        r"(\d{4,10})([^\[\(]*)(?:\[([^]]+)\]|\([^\)]*\))?",
        tbody_text,
        flags=re.UNICODE,
    )

    refmatches = re.findall(
        # r"\b(\d{4,10})*(?:\[([^\]]+)\]|\(([^)]+)\))?([a-zA-Z]+)?",
        r"\b(\d{4,6})([a-zA-Z]+)?\b",
        tbody_text,
        flags=re.UNICODE,
    )
    testlog = extract_brackets(tbody_text)
    coler_text_array = []
    for log in testlog:
        (log[0] if log[0] else log[1])

        coler_text_array.append(str(log[0]) if log[0] else str(log[1]))

    save_logs_to_file(coler_text_array, "kakolog.txt")

    #!テストここまで

    #!リファレンスナンバー配列取得
    # refarray = ref_array_make(refmatches)

    # 色
    # colormatches = re.findall(r"\[([^\]]+)\]|\(([^)]+)\)", tbody_text, flags=re.UNICODE)
    # colormatches = re.findall(r"\[([^\]]+)\]|\((.+?)\)", tbody_text, flags=re.UNICODE)
    # colormatches = re.findall(r"\[([^\]]+)\]|\(([^)]+)\)", tbody_text, flags=re.UNICODE)
    # colormatches =  re.findall(r'\[([^\]]+)\]|\((.*?)\)', tbody_text)

    # colormatches = re.findall(r'[^\[\(]*\[(([^\]]+))\]|[^\[\(]*\(((.*?))\)', tbody_text)
    # colormatches = re.findall(r'\((.*?)\)お気に入り登録', tbody_text)
    # colormatches = re.findall(r"\b(\d{4,6})([a-zA-Z]+)?\s*\((.*?)\)お気に入り登録¥(\d{1,3}(?:,\d{3})*|―)",tbody_text)

    # colormatches = re.findall(r"\b(\d{4,6})([a-zA-Z]+)?\s*\((.*?)\)お気に入り登録.*?¥(\d{1,3}(,\d{3})*)",tbody_text)

    # colormatches = re.findall(r"\b(\d{4,6})([a-zA-Z]+)?\s*\((.*?)\)お気に入り登録.*?¥(\d{1,3}(?:,\d{3})*|―)",tbody_text)
    # colormatches = re.findall(r"\b(\d{4,6})([a-zA-Z]+)?\s*\((.*?)\)お気に入り登録.*?¥(\d{1,3}(?:,\d{3}|\d{3})*|―)", tbody_text)

    # colormatches = re.findall(r"\b(\d{4,6})([a-zA-Z]+)?\s*\((.*?)\)お気に入り登録.*?¥(\d{1,3}(?:,\d{3})*|―)", tbody_text)

    colormatches = re.findall(
        r"\b(\d{4,6})([a-zA-Z]+)?\s*\((.*?)\)お気に入り登録.*?¥(\d{1,3}(?:,\d{3})*|―).*?¥(\d{1,3}(?:,\d{3})*|―)",
        tbody_text,
    )
    colormatches2 = re.findall(
        r"\b(\d{4,6})([a-zA-Z]+)?\s*\[(.*?)\]お気に入り登録.*?¥(\d{1,3}(?:,\d{3})*|―).*?¥(\d{1,3}(?:,\d{3})*|―)",
        tbody_text,
    )

    rankingpattern = re.findall(r"(\d+)位(\d+)位", tbody_text)

    # !色とアイテム名配列取得
    # colorarray = color_array_make(colormatches)
    colorarray = color_array_make(colormatches)
    colorarray2 = color_array_make(colormatches2)
    result_list = list(
        map(lambda item: "".join(re.findall(r"'(.*?)'", item)[:2]), colorarray2)
    )
    for item in result_list:
        (item)
    exceldatas = []
    index = 0
    for item in colorarray:
        index += 1
        itemmatches = re.findall(r"'(.*?)'", item)
        combined_element = itemmatches[0] + itemmatches[1]
        # 0番目と1番目の要素を削除
        del itemmatches[0]
        del itemmatches[0]
        itemmatches.insert(0, combined_element)
        ws.append(itemmatches)
        # ws[f"A{index}"] = itemmatches[0]+itemmatches[1]
        # ws[f"B{index}"] = itemmatches[2]
        # ws[f"C{index}"] = itemmatches[3]
        # ws[f"D{index}"] = itemmatches[4]
        (itemmatches[0])

    index = 0
    for item in colorarray2:
        index += 1
        itemmatches = re.findall(r"'(.*?)'", item)
        combined_element = itemmatches[0] + itemmatches[1]
        # 0番目と1番目の要素を削除
        del itemmatches[0]
        del itemmatches[0]
        itemmatches.insert(0, combined_element)
        ws.append(itemmatches)
        # ws[f"A{index}"] = itemmatches[0]+itemmatches[1]
        # ws[f"B{index}"] = itemmatches[2]
        # ws[f"C{index}"] = itemmatches[3]
        # ws[f"D{index}"] = itemmatches[4]
        (itemmatches[0])
    # colorarray = color_array_make(text_words)
    refandcollor_array_make(refandcolormatches)
    # save_logs_to_file(parenthesesvaluematches, "reflogfile.txt")
    # for index , refandcolormatche in enumerate(refandcolormatches):

    # 以下金額の正規表現
    pricevaluematches = re.findall(r"¥(\d{1,3}(?:,\d{3})*|―)", tbody_text)
    # 二次元配列が返ってくる。最安値と最高値
    # prices_array_make = refanditemname
    # !金額配列取得(最安値と高値を二次元配列に格納している)
    pricearray = prices_array_make(pricevaluematches)
    # pricevaluematches = re.findall(r"￥-?(\d{1,3}(?:,\d{3})*)", tbody_text)
    # pricevaluematches = re.findall(r"￥-?(\d{0,3}(?:,\d{3})*)", tbody_text)
    # pricevaluematches = re.findall(r"￥(\d{1,3}(?:,\d{3})*)|￥-?", tbody_text)
    # pricevaluematches = re.findall(r"￥(\d{1,3}(?:,\d{3})*)|-", tbody_text)
    # !モデル
    # !model_info = re.search(r'ロレックス.*?\d{6}[A-Z]+', text).group()  # ロレックスGMTマスターII 126710BLNR
    parenthesesvalues = []
    parenthindex = 0

    # ログをファイルに保存
    save_logs_to_file(pricevaluematches, "log_file.txt")
    # save_logs_to_file(tbody_text, "log_file2.txt")
    # 金額抽出ループ
    # 金額抽出ループ
    itemlist = [None] * len(pricevaluematches)

    # !colorarrayが正しい要素数なので、それにあわせて配列を合成する
    (f"{len(colorarray)}はアイテム名の数です")
    (f"{len(pricearray)}は金額の数です")
    itemsarray = []

    if len(colorarray) < 40:
        item_index = len(colorarray)
    elif len(colorarray) >= 40:
        item_index = 40
    for i in range(item_index):
        index = i + 1
        # ws[f"A{index}"] = colorarray[i]
        # ws[f"B{index}"] = refarray[i]
        # ws[f"C{index}"] = pricearray[i][0]
        # ws[f"D{index}"] = pricearray[i][1]

        # ws.append([colorarray[index],refarray[index],pricearray[[index][0]],pricearray[[index][1]]])

    for index, pricematch in enumerate(pricevaluematches):
        # itemlist[index].insert(0, pricematch)
        (f"{pricematch}は金額です!!")

    # かっこに囲まれた文字列をループする。
    for index, parenthesevaluematch in enumerate(parenthesesvaluematches):
        # itemlist.insert[index[1], parenthesevaluematch]
        # parenthesesvalues.append(parenthesevaluematch)
        (f"{parenthesevaluematch}はかっこに囲まれた値")
    itemlistindex = 0
    save_logs_to_file(tbody_text, "tbody.txt")
    for index, word in enumerate(text_words):
        # itemlistindex += 1
        # itemlist.append(word)
        ("------------------ここで分割される----------------")
        if price_validate_imput(word):
            (f"これは金額です:{word}")
        if validate_input(word):
            # trueなら次の製品にいき、かつ、配列を空にする
            (f"{word}")

            ("------------------ここでアイテムごとの処理は終了!!----------------")
            itemlistindex += 1
            # itemlist.insert(index[2], word)
            # itemlist.append(word)
            # (parenthesesvalues[parenthindex])
            # 上記で追加した商品名を既存配列に追加する。
            # new_values = [parenthesesvalues[parenthindex]] + itemlist[:-1]
            # 特定の列にデータを入稿する。
            # ws.append(itemlist)
            # ws.append([1, 2, 3, word])
            # itemlist = []
            parenthindex += 1
            # itemlistindex = 0
            # データ入稿してcontinue
            continue
        else:
            (word)
            # itemlist.append(word)
        # if "ロレックス" in word:
        #     (f"テキスト中に 'ロレックス' が含まれています: {word}")
    (f"{itemlistindex}は必要データ数です")
    # 改行で分割してリストにする

    text_lines = tbody_text.split("\n")

    # # 各行に対してループを行い、"ロレックス" を含むかどうかを判定
    # for line in text_lines:
    #     (f"{line}---------")
    # if "ロレックス" in line:
    #     (f"テキスト中に 'ロレックス' が含まれています: {line}")

    all_elements = tbody_tag.find_all(True, recursive=False)

    # 各要素から情報を抽出して表示
    for element in all_elements:
        tag_name = element.name
        class_name = element.get("class", None)
        text_content = element.get_text(strip=True)

        # (f"Tag: {tag_name}, Class: {class_name}, Text: {text_content}")

else:
    ("<tbody> タグが見つかりませんでした。")
# エクセルファイルを保存
wb.save(file_name)
# WebDriverを終了
driver.quit()
